from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Query
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import httpx
import uuid
import aiofiles
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
from datetime import datetime, timezone, timedelta

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Static files directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ─── Pydantic Models ───

class UserOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    picture: Optional[str] = None
    role: Optional[str] = "user"  # user, moderator, admin
    created_at: Optional[str] = None

# Moderator emails
MODERATOR_EMAILS = ["topkitfr@gmail.com"]

class MasterKitCreate(BaseModel):
    club: str
    season: str
    kit_type: str  # Home/Away/Third/Fourth/GK/Special/Other
    brand: str
    front_photo: str
    league: Optional[str] = ""
    design: Optional[str] = ""
    sponsor: Optional[str] = ""
    gender: Optional[str] = ""  # Man/Woman/Kid

class MasterKitOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    kit_id: str
    club: str
    season: str
    kit_type: str
    brand: str
    front_photo: str
    league: Optional[str] = ""
    design: Optional[str] = ""
    sponsor: Optional[str] = ""
    gender: Optional[str] = ""
    created_by: Optional[str] = None
    created_at: Optional[str] = None
    version_count: Optional[int] = 0
    avg_rating: Optional[float] = 0.0

class VersionCreate(BaseModel):
    kit_id: str
    competition: str  # National Championship/National Cup/Continental Cup/Intercontinental Cup/World Cup
    model: str  # Authentic/Replica/Other
    sku_code: Optional[str] = ""
    ean_code: Optional[str] = ""
    front_photo: Optional[str] = ""
    back_photo: Optional[str] = ""

class VersionOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    version_id: str
    kit_id: str
    competition: str
    model: str
    sku_code: Optional[str] = ""
    ean_code: Optional[str] = ""
    front_photo: Optional[str] = ""
    back_photo: Optional[str] = ""
    created_by: Optional[str] = None
    created_at: Optional[str] = None
    avg_rating: Optional[float] = 0.0
    review_count: Optional[int] = 0

class CollectionAdd(BaseModel):
    version_id: str
    category: Optional[str] = "General"
    notes: Optional[str] = ""
    # Flocking
    flocking_type: Optional[str] = ""  # Name+Number / Name / Number
    flocking_origin: Optional[str] = ""  # Official / Personalized
    flocking_detail: Optional[str] = ""  # e.g., "Messi 10"
    # Condition
    condition_origin: Optional[str] = ""  # Club Stock / Match Prepared / Match Worn / Training / Shop
    physical_state: Optional[str] = ""  # New with tag / Very good / Used / Damaged / Needs restoration
    size: Optional[str] = ""
    # Values
    purchase_cost: Optional[float] = None
    estimated_price: Optional[float] = None
    price_estimate: Optional[float] = None
    value_estimate: Optional[float] = None
    # Signed
    signed: Optional[bool] = False
    signed_by: Optional[str] = ""
    signed_proof: Optional[bool] = False
    # Legacy
    condition: Optional[str] = ""
    printing: Optional[str] = ""

class CollectionUpdate(BaseModel):
    category: Optional[str] = None
    notes: Optional[str] = None
    flocking_type: Optional[str] = None
    flocking_origin: Optional[str] = None
    flocking_detail: Optional[str] = None
    condition_origin: Optional[str] = None
    physical_state: Optional[str] = None
    size: Optional[str] = None
    purchase_cost: Optional[float] = None
    estimated_price: Optional[float] = None
    price_estimate: Optional[float] = None
    value_estimate: Optional[float] = None
    signed: Optional[bool] = None
    signed_by: Optional[str] = None
    signed_proof: Optional[bool] = None
    condition: Optional[str] = None
    printing: Optional[str] = None

class CollectionOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    collection_id: str
    user_id: str
    version_id: str
    category: str
    notes: Optional[str] = ""
    flocking_type: Optional[str] = ""
    flocking_origin: Optional[str] = ""
    flocking_detail: Optional[str] = ""
    condition_origin: Optional[str] = ""
    physical_state: Optional[str] = ""
    size: Optional[str] = ""
    purchase_cost: Optional[float] = None
    estimated_price: Optional[float] = None
    price_estimate: Optional[float] = None
    value_estimate: Optional[float] = None
    signed: Optional[bool] = False
    signed_by: Optional[str] = ""
    signed_proof: Optional[bool] = False
    condition: Optional[str] = ""
    printing: Optional[str] = ""
    added_at: Optional[str] = None

class ReviewCreate(BaseModel):
    version_id: str
    rating: int  # 1-5
    comment: Optional[str] = ""

class ReviewOut(BaseModel):
    model_config = ConfigDict(extra="ignore")
    review_id: str
    version_id: str
    user_id: str
    user_name: Optional[str] = ""
    user_picture: Optional[str] = ""
    rating: int
    comment: Optional[str] = ""
    created_at: Optional[str] = None

class SubmissionCreate(BaseModel):
    submission_type: str  # master_kit or version
    data: dict

class VoteCreate(BaseModel):
    vote: str  # up or down

class ReportCreate(BaseModel):
    target_type: str  # master_kit or version
    target_id: str
    corrections: dict
    notes: Optional[str] = ""

class ProfileUpdate(BaseModel):
    username: Optional[str] = None
    description: Optional[str] = None
    collection_privacy: Optional[str] = None
    profile_picture: Optional[str] = None

APPROVAL_THRESHOLD = 5
MODERATOR_APPROVAL_THRESHOLD = 1  # Moderators can approve with single vote


# ─── Auth Helpers ───

EMERGENT_AUTH_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

async def get_current_user(request: Request) -> dict:
    session_token = request.cookies.get("session_token")
    if not session_token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            session_token = auth_header[7:]
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session_doc = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
    if not session_doc:
        raise HTTPException(status_code=401, detail="Invalid session")
    expires_at = session_doc["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Session expired")
    user_doc = await db.users.find_one({"user_id": session_doc["user_id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="User not found")
    return user_doc


# ─── Auth Routes ───

@api_router.post("/auth/session")
async def create_session(request: Request, response: Response):
    body = await request.json()
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    async with httpx.AsyncClient() as hc:
        resp = await hc.get(EMERGENT_AUTH_URL, headers={"X-Session-ID": session_id})
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session_id")
        data = resp.json()
    email = data["email"]
    name = data.get("name", "")
    picture = data.get("picture", "")
    session_token = data["session_token"]
    # Determine role based on email
    role = "moderator" if email in MODERATOR_EMAILS else "user"
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one({"email": email}, {"$set": {"name": name, "picture": picture, "role": role}})
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": name,
            "picture": picture,
            "role": role,
            "username": name.replace(" ", "").lower() if name else "",
            "description": "",
            "collection_privacy": "public",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    response.set_cookie(
        key="session_token", value=session_token,
        httponly=True, secure=True, samesite="none",
        path="/", max_age=7*24*60*60
    )
    user_doc = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return user_doc

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    # Ensure role is included
    if "role" not in user:
        user["role"] = "moderator" if user.get("email") in MODERATOR_EMAILS else "user"
    return user

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie(key="session_token", path="/", secure=True, samesite="none")
    return {"message": "Logged out"}


# ─── Master Kit Routes ───

@api_router.get("/master-kits", response_model=List[MasterKitOut])
async def list_master_kits(
    club: Optional[str] = None,
    season: Optional[str] = None,
    brand: Optional[str] = None,
    kit_type: Optional[str] = None,
    design: Optional[str] = None,
    league: Optional[str] = None,
    gender: Optional[str] = None,
    search: Optional[str] = None,
    skip: int = 0,
    limit: int = 200
):
    query = {}
    if club:
        query["club"] = {"$regex": club, "$options": "i"}
    if season:
        query["season"] = {"$regex": season, "$options": "i"}
    if brand:
        query["brand"] = {"$regex": brand, "$options": "i"}
    if kit_type:
        query["kit_type"] = kit_type
    if design:
        query["design"] = design
    if league:
        query["league"] = league
    if gender:
        query["gender"] = gender
    if search:
        query["$or"] = [
            {"club": {"$regex": search, "$options": "i"}},
            {"brand": {"$regex": search, "$options": "i"}},
            {"season": {"$regex": search, "$options": "i"}},
            {"design": {"$regex": search, "$options": "i"}},
            {"sponsor": {"$regex": search, "$options": "i"}},
        ]
    kits = await db.master_kits.find(query, {"_id": 0}).sort("season", -1).skip(skip).limit(limit).to_list(limit)
    for kit in kits:
        version_count = await db.versions.count_documents({"kit_id": kit["kit_id"]})
        kit["version_count"] = version_count
        reviews = await db.reviews.find({"kit_id": kit["kit_id"]}, {"_id": 0, "rating": 1}).to_list(1000)
        if reviews:
            kit["avg_rating"] = round(sum(r["rating"] for r in reviews) / len(reviews), 1)
        else:
            kit["avg_rating"] = 0.0
    return kits

@api_router.get("/master-kits/count")
async def count_master_kits():
    count = await db.master_kits.count_documents({})
    return {"count": count}

@api_router.get("/master-kits/filters")
async def get_filters():
    clubs = await db.master_kits.distinct("club")
    brands = await db.master_kits.distinct("brand")
    seasons = await db.master_kits.distinct("season")
    kit_types = await db.master_kits.distinct("kit_type")
    designs = await db.master_kits.distinct("design")
    leagues = await db.master_kits.distinct("league")
    sponsors = await db.master_kits.distinct("sponsor")
    genders = await db.master_kits.distinct("gender")
    return {
        "clubs": sorted([c for c in clubs if c]),
        "brands": sorted([b for b in brands if b]),
        "seasons": sorted([s for s in seasons if s], reverse=True),
        "kit_types": sorted([t for t in kit_types if t]),
        "designs": sorted([d for d in designs if d]),
        "leagues": sorted([lg for lg in leagues if lg]),
        "sponsors": sorted([s for s in sponsors if s]),
        "genders": sorted([g for g in genders if g]),
    }

@api_router.get("/master-kits/{kit_id}")
async def get_master_kit(kit_id: str):
    kit = await db.master_kits.find_one({"kit_id": kit_id}, {"_id": 0})
    if not kit:
        raise HTTPException(status_code=404, detail="Kit not found")
    versions = await db.versions.find({"kit_id": kit_id}, {"_id": 0}).to_list(100)
    for v in versions:
        reviews = await db.reviews.find({"version_id": v["version_id"]}, {"_id": 0, "rating": 1}).to_list(1000)
        v["avg_rating"] = round(sum(r["rating"] for r in reviews) / len(reviews), 1) if reviews else 0.0
        v["review_count"] = len(reviews)
    kit["versions"] = versions
    kit["version_count"] = len(versions)
    all_reviews = await db.reviews.find({"kit_id": kit_id}, {"_id": 0, "rating": 1}).to_list(1000)
    kit["avg_rating"] = round(sum(r["rating"] for r in all_reviews) / len(all_reviews), 1) if all_reviews else 0.0
    return kit

@api_router.post("/master-kits", response_model=MasterKitOut)
async def create_master_kit(kit: MasterKitCreate, request: Request):
    user = await get_current_user(request)
    doc = kit.model_dump()
    doc["kit_id"] = f"kit_{uuid.uuid4().hex[:12]}"
    doc["created_by"] = user["user_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.master_kits.insert_one(doc)
    
    # Auto-create a default Version for this Master Kit
    default_version = {
        "version_id": f"ver_{uuid.uuid4().hex[:12]}",
        "kit_id": doc["kit_id"],
        "competition": "National Championship",
        "model": "Replica",
        "sku_code": "",
        "ean_code": "",
        "front_photo": doc.get("front_photo", ""),
        "back_photo": "",
        "created_by": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.versions.insert_one(default_version)
    
    result = await db.master_kits.find_one({"kit_id": doc["kit_id"]}, {"_id": 0})
    result["version_count"] = 1
    result["avg_rating"] = 0.0
    return result


# ─── Version Routes ───

@api_router.get("/versions/{version_id}")
async def get_version(version_id: str):
    version = await db.versions.find_one({"version_id": version_id}, {"_id": 0})
    if not version:
        raise HTTPException(status_code=404, detail="Version not found")
    kit = await db.master_kits.find_one({"kit_id": version["kit_id"]}, {"_id": 0})
    version["master_kit"] = kit
    reviews = await db.reviews.find({"version_id": version_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for r in reviews:
        u = await db.users.find_one({"user_id": r["user_id"]}, {"_id": 0, "name": 1, "picture": 1})
        if u:
            r["user_name"] = u.get("name", "")
            r["user_picture"] = u.get("picture", "")
    version["reviews"] = reviews
    version["avg_rating"] = round(sum(r["rating"] for r in reviews) / len(reviews), 1) if reviews else 0.0
    version["review_count"] = len(reviews)
    collection_count = await db.collections.count_documents({"version_id": version_id})
    version["collection_count"] = collection_count
    return version

@api_router.get("/versions/{version_id}/estimates")
async def get_version_estimates(version_id: str):
    items = await db.collections.find(
        {"version_id": version_id, "value_estimate": {"$gt": 0}},
        {"_id": 0, "value_estimate": 1}
    ).to_list(1000)
    estimates = [i["value_estimate"] for i in items if i.get("value_estimate")]
    if not estimates:
        return {"low": 0, "average": 0, "high": 0, "count": 0, "estimates": []}
    return {
        "low": round(min(estimates), 2),
        "average": round(sum(estimates) / len(estimates), 2),
        "high": round(max(estimates), 2),
        "count": len(estimates),
        "estimates": sorted(estimates)
    }

@api_router.post("/versions", response_model=VersionOut)
async def create_version(version: VersionCreate, request: Request):
    user = await get_current_user(request)
    kit = await db.master_kits.find_one({"kit_id": version.kit_id}, {"_id": 0})
    if not kit:
        raise HTTPException(status_code=404, detail="Master Kit not found")
    doc = version.model_dump()
    doc["version_id"] = f"ver_{uuid.uuid4().hex[:12]}"
    doc["created_by"] = user["user_id"]
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    if not doc["front_photo"]:
        doc["front_photo"] = kit.get("front_photo", "")
    await db.versions.insert_one(doc)
    result = await db.versions.find_one({"version_id": doc["version_id"]}, {"_id": 0})
    result["avg_rating"] = 0.0
    result["review_count"] = 0
    return result

@api_router.get("/versions", response_model=List[VersionOut])
async def list_versions(kit_id: Optional[str] = None, skip: int = 0, limit: int = 50):
    query = {}
    if kit_id:
        query["kit_id"] = kit_id
    versions = await db.versions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    for v in versions:
        reviews = await db.reviews.find({"version_id": v["version_id"]}, {"_id": 0, "rating": 1}).to_list(1000)
        v["avg_rating"] = round(sum(r["rating"] for r in reviews) / len(reviews), 1) if reviews else 0.0
        v["review_count"] = len(reviews)
    return versions


# ─── Collection Routes ───

@api_router.get("/collections")
async def get_my_collection(request: Request, category: Optional[str] = None):
    user = await get_current_user(request)
    query = {"user_id": user["user_id"]}
    if category:
        query["category"] = category
    items = await db.collections.find(query, {"_id": 0}).sort("added_at", -1).to_list(500)
    enriched = []
    for item in items:
        version = await db.versions.find_one({"version_id": item["version_id"]}, {"_id": 0})
        if version:
            kit = await db.master_kits.find_one({"kit_id": version["kit_id"]}, {"_id": 0})
            item["version"] = version
            item["master_kit"] = kit
        enriched.append(item)
    return enriched

@api_router.get("/collections/categories")
async def get_collection_categories(request: Request):
    user = await get_current_user(request)
    cats = await db.collections.distinct("category", {"user_id": user["user_id"]})
    return sorted(cats)

@api_router.post("/collections")
async def add_to_collection(item: CollectionAdd, request: Request):
    user = await get_current_user(request)
    version = await db.versions.find_one({"version_id": item.version_id}, {"_id": 0})
    if not version:
        raise HTTPException(status_code=404, detail="Version not found")
    existing = await db.collections.find_one(
        {"user_id": user["user_id"], "version_id": item.version_id}, {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Already in collection")
    doc = {
        "collection_id": f"col_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "version_id": item.version_id,
        "category": item.category or "General",
        "notes": item.notes or "",
        "flocking_type": item.flocking_type or "",
        "flocking_origin": item.flocking_origin or "",
        "flocking_detail": item.flocking_detail or "",
        "condition_origin": item.condition_origin or "",
        "physical_state": item.physical_state or "",
        "size": item.size or "",
        "purchase_cost": item.purchase_cost,
        "estimated_price": item.estimated_price,
        "price_estimate": item.price_estimate,
        "value_estimate": item.value_estimate,
        "signed": item.signed or False,
        "signed_by": item.signed_by or "",
        "signed_proof": item.signed_proof or False,
        "condition": item.condition or "",
        "printing": item.printing or "",
        "added_at": datetime.now(timezone.utc).isoformat()
    }
    await db.collections.insert_one(doc)
    result = await db.collections.find_one({"collection_id": doc["collection_id"]}, {"_id": 0})
    return result

@api_router.delete("/collections/{collection_id}")
async def remove_from_collection(collection_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.collections.delete_one({"collection_id": collection_id, "user_id": user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Removed from collection"}

@api_router.put("/collections/{collection_id}")
async def update_collection_item(collection_id: str, update: CollectionUpdate, request: Request):
    user = await get_current_user(request)
    update_dict = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    result = await db.collections.update_one(
        {"collection_id": collection_id, "user_id": user["user_id"]},
        {"$set": update_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    updated = await db.collections.find_one({"collection_id": collection_id}, {"_id": 0})
    return updated

@api_router.get("/collections/stats")
async def get_collection_stats(request: Request):
    user = await get_current_user(request)
    items = await db.collections.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(1000)
    total = len(items)
    estimates = [i["value_estimate"] for i in items if i.get("value_estimate") and i["value_estimate"] > 0]
    low = min(estimates) * total if estimates else 0
    avg = sum(estimates) if estimates else 0
    high = max(estimates) * total if estimates else 0
    return {
        "total_jerseys": total,
        "estimated_value": {"low": round(low, 2), "average": round(avg, 2), "high": round(high, 2)},
        "items_with_estimates": len(estimates)
    }

@api_router.get("/collections/category-stats")
async def get_category_stats(request: Request):
    user = await get_current_user(request)
    items = await db.collections.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(1000)
    categories = {}
    for item in items:
        cat = item.get("category", "General")
        if cat not in categories:
            categories[cat] = {"count": 0, "estimates": []}
        categories[cat]["count"] += 1
        if item.get("value_estimate") and item["value_estimate"] > 0:
            categories[cat]["estimates"].append(item["value_estimate"])
    result = []
    for cat, data in categories.items():
        est = data["estimates"]
        result.append({
            "category": cat,
            "count": data["count"],
            "estimated_value": {
                "low": round(min(est), 2) if est else 0,
                "average": round(sum(est) / len(est), 2) if est else 0,
                "high": round(max(est), 2) if est else 0,
            }
        })
    return result


# ─── Estimation Logic ───

ESTIMATION_BASE_PRICES = {"Authentic": 140, "Replica": 90, "Other": 60}
ESTIMATION_COMPETITION_COEFF = {
    "National Championship": 0.0,
    "National Cup": 0.05,
    "Continental Cup": 1.0,
    "Intercontinental Cup": 1.0,
    "World Cup": 1.0,
}
ESTIMATION_ORIGIN_COEFF = {"Club Stock": 0.5, "Match Prepared": 1.0, "Match Worn": 1.5, "Training": 0.0, "Shop": 0.0}
ESTIMATION_STATE_COEFF = {"New with tag": 0.3, "Very good": 0.1, "Used": 0.0, "Damaged": -0.2, "Needs restoration": -0.4}
ESTIMATION_FLOCKING_COEFF = {"Official": 0.15, "Personalized": 0.0}
ESTIMATION_SIGNED_COEFF = 1.5
ESTIMATION_SIGNED_PROOF_COEFF = 1.0
ESTIMATION_AGE_COEFF_PER_YEAR = 0.05
ESTIMATION_AGE_MAX = 1.0

def calculate_estimation(model_type: str, competition: str, condition_origin: str,
                         physical_state: str, flocking_origin: str,
                         signed: bool, signed_proof: bool, season_year: int):
    base = ESTIMATION_BASE_PRICES.get(model_type, 60)
    coeff_sum = 0.0
    breakdown = []

    # Competition coefficient
    comp_c = ESTIMATION_COMPETITION_COEFF.get(competition, 0.0)
    coeff_sum += comp_c
    if competition:
        breakdown.append({"label": f"Competition: {competition}", "coeff": comp_c})

    # Origin coefficient
    origin_c = ESTIMATION_ORIGIN_COEFF.get(condition_origin, 0.0)
    coeff_sum += origin_c
    if condition_origin:
        breakdown.append({"label": f"Origin: {condition_origin}", "coeff": origin_c})

    # Physical state coefficient
    state_c = ESTIMATION_STATE_COEFF.get(physical_state, 0.0)
    coeff_sum += state_c
    if physical_state:
        breakdown.append({"label": f"State: {physical_state}", "coeff": state_c})

    # Flocking coefficient
    flocking_c = ESTIMATION_FLOCKING_COEFF.get(flocking_origin, 0.0)
    coeff_sum += flocking_c
    if flocking_origin:
        breakdown.append({"label": f"Flocking: {flocking_origin}", "coeff": flocking_c})

    # Signed coefficient
    if signed:
        coeff_sum += ESTIMATION_SIGNED_COEFF
        breakdown.append({"label": "Signed", "coeff": ESTIMATION_SIGNED_COEFF})
        if signed_proof:
            coeff_sum += ESTIMATION_SIGNED_PROOF_COEFF
            breakdown.append({"label": "Proof/Certificate", "coeff": ESTIMATION_SIGNED_PROOF_COEFF})

    # Age coefficient
    current_year = datetime.now(timezone.utc).year
    age = max(0, current_year - season_year) if season_year else 0
    age_c = min(age * ESTIMATION_AGE_COEFF_PER_YEAR, ESTIMATION_AGE_MAX)
    coeff_sum += age_c
    if age > 0:
        breakdown.append({"label": f"Age: {age} years", "coeff": round(age_c, 2)})

    estimated_price = round(base * (1 + coeff_sum), 2)
    return {
        "base_price": base,
        "model_type": model_type,
        "coeff_sum": round(coeff_sum, 2),
        "estimated_price": estimated_price,
        "breakdown": breakdown
    }

class EstimationRequest(BaseModel):
    model_type: str  # Authentic / Replica / Other
    competition: Optional[str] = ""
    condition_origin: Optional[str] = ""
    physical_state: Optional[str] = ""
    flocking_origin: Optional[str] = ""
    signed: Optional[bool] = False
    signed_proof: Optional[bool] = False
    season_year: Optional[int] = 0

@api_router.post("/estimate")
async def estimate_price(req: EstimationRequest):
    result = calculate_estimation(
        model_type=req.model_type,
        competition=req.competition or "",
        condition_origin=req.condition_origin or "",
        physical_state=req.physical_state or "",
        flocking_origin=req.flocking_origin or "",
        signed=req.signed or False,
        signed_proof=req.signed_proof or False,
        season_year=req.season_year or 0
    )
    return result


# ─── Review Routes ───

@api_router.post("/reviews")
async def create_review(review: ReviewCreate, request: Request):
    user = await get_current_user(request)
    if review.rating < 1 or review.rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be 1-5")
    version = await db.versions.find_one({"version_id": review.version_id}, {"_id": 0})
    if not version:
        raise HTTPException(status_code=404, detail="Version not found")
    existing = await db.reviews.find_one(
        {"user_id": user["user_id"], "version_id": review.version_id}, {"_id": 0}
    )
    if existing:
        await db.reviews.update_one(
            {"review_id": existing["review_id"]},
            {"$set": {"rating": review.rating, "comment": review.comment, "created_at": datetime.now(timezone.utc).isoformat()}}
        )
        updated = await db.reviews.find_one({"review_id": existing["review_id"]}, {"_id": 0})
        return updated
    doc = {
        "review_id": f"rev_{uuid.uuid4().hex[:12]}",
        "version_id": review.version_id,
        "kit_id": version["kit_id"],
        "user_id": user["user_id"],
        "user_name": user.get("name", ""),
        "user_picture": user.get("picture", ""),
        "rating": review.rating,
        "comment": review.comment or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.reviews.insert_one(doc)
    result = await db.reviews.find_one({"review_id": doc["review_id"]}, {"_id": 0})
    return result

@api_router.get("/reviews")
async def get_reviews(version_id: str):
    reviews = await db.reviews.find({"version_id": version_id}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for r in reviews:
        u = await db.users.find_one({"user_id": r["user_id"]}, {"_id": 0, "name": 1, "picture": 1})
        if u:
            r["user_name"] = u.get("name", "")
            r["user_picture"] = u.get("picture", "")
    return reviews


# ─── Stats Routes ───

@api_router.get("/stats")
async def get_stats():
    kits = await db.master_kits.count_documents({})
    versions = await db.versions.count_documents({})
    users = await db.users.count_documents({})
    reviews = await db.reviews.count_documents({})
    return {"master_kits": kits, "versions": versions, "users": users, "reviews": reviews}


# ─── User Profile ───

@api_router.get("/users/{user_id}/profile")
async def get_user_profile(user_id: str):
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    collection_count = await db.collections.count_documents({"user_id": user_id})
    review_count = await db.reviews.count_documents({"user_id": user_id})
    submission_count = await db.submissions.count_documents({"submitted_by": user_id})
    user["collection_count"] = collection_count
    user["review_count"] = review_count
    user["submission_count"] = submission_count
    return user

@api_router.put("/users/profile")
async def update_profile(update: ProfileUpdate, request: Request):
    user = await get_current_user(request)
    update_dict = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    if "username" in update_dict:
        existing = await db.users.find_one(
            {"username": update_dict["username"], "user_id": {"$ne": user["user_id"]}}, {"_id": 0}
        )
        if existing:
            raise HTTPException(status_code=400, detail="Username already taken")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": update_dict})
    updated = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    return updated


# ─── Submission Routes ───

@api_router.post("/submissions")
async def create_submission(sub: SubmissionCreate, request: Request):
    user = await get_current_user(request)
    if sub.submission_type not in ("master_kit", "version"):
        raise HTTPException(status_code=400, detail="Invalid submission type")
    doc = {
        "submission_id": f"sub_{uuid.uuid4().hex[:12]}",
        "submission_type": sub.submission_type,
        "data": sub.data,
        "submitted_by": user["user_id"],
        "submitter_name": user.get("name", ""),
        "status": "pending",
        "votes_up": 0,
        "votes_down": 0,
        "voters": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.submissions.insert_one(doc)
    result = await db.submissions.find_one({"submission_id": doc["submission_id"]}, {"_id": 0})
    return result

@api_router.get("/submissions")
async def list_submissions(status: Optional[str] = "pending", skip: int = 0, limit: int = 50):
    query = {}
    if status:
        query["status"] = status
    subs = await db.submissions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return subs

@api_router.get("/submissions/{submission_id}")
async def get_submission(submission_id: str):
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    return sub

@api_router.post("/submissions/{submission_id}/vote")
async def vote_on_submission(submission_id: str, vote: VoteCreate, request: Request):
    user = await get_current_user(request)
    # Check user has at least 1 jersey in collection (moderators exempt)
    user_role = user.get("role", "user")
    is_moderator = user_role in ("moderator", "admin")
    if not is_moderator:
        col_count = await db.collections.count_documents({"user_id": user["user_id"]})
        if col_count == 0:
            raise HTTPException(status_code=403, detail="You must have at least 1 jersey in your collection to vote")
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(status_code=404, detail="Submission not found")
    if sub["status"] != "pending":
        raise HTTPException(status_code=400, detail="Submission is no longer pending")
    if user["user_id"] in sub.get("voters", []):
        raise HTTPException(status_code=400, detail="Already voted")
    if vote.vote not in ("up", "down"):
        raise HTTPException(status_code=400, detail="Vote must be 'up' or 'down'")
    
    # Moderators can approve with single vote
    vote_weight = APPROVAL_THRESHOLD if is_moderator and vote.vote == "up" else 1
    inc_field = "votes_up" if vote.vote == "up" else "votes_down"
    await db.submissions.update_one(
        {"submission_id": submission_id},
        {"$inc": {inc_field: vote_weight}, "$push": {"voters": user["user_id"]}}
    )
    # Check if threshold reached
    updated_sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if updated_sub["votes_up"] >= APPROVAL_THRESHOLD:
        # Approve: create the actual kit/version
        data = updated_sub["data"]
        if updated_sub["submission_type"] == "master_kit":
            kit_id = f"kit_{uuid.uuid4().hex[:12]}"
            kit_doc = {
                "kit_id": kit_id,
                "club": data.get("club", ""),
                "season": data.get("season", ""),
                "kit_type": data.get("kit_type", ""),
                "brand": data.get("brand", ""),
                "front_photo": data.get("front_photo", ""),
                "league": data.get("league", ""),
                "design": data.get("design", ""),
                "sponsor": data.get("sponsor", ""),
                "gender": data.get("gender", ""),
                "created_by": updated_sub["submitted_by"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.master_kits.insert_one(kit_doc)
            # Auto-create a default Version for this Master Kit
            default_version = {
                "version_id": f"ver_{uuid.uuid4().hex[:12]}",
                "kit_id": kit_id,
                "competition": "National Championship",
                "model": "Replica",
                "sku_code": "",
                "ean_code": "",
                "front_photo": data.get("front_photo", ""),
                "back_photo": "",
                "created_by": updated_sub["submitted_by"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.versions.insert_one(default_version)
        elif updated_sub["submission_type"] == "version":
            ver_doc = {
                "version_id": f"ver_{uuid.uuid4().hex[:12]}",
                "kit_id": data.get("kit_id", ""),
                "competition": data.get("competition", ""),
                "model": data.get("model", ""),
                "sku_code": data.get("sku_code", ""),
                "ean_code": data.get("ean_code", ""),
                "front_photo": data.get("front_photo", ""),
                "back_photo": data.get("back_photo", ""),
                "created_by": updated_sub["submitted_by"],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.versions.insert_one(ver_doc)
        await db.submissions.update_one(
            {"submission_id": submission_id},
            {"$set": {"status": "approved"}}
        )
    return await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})


# ─── Report Routes ───

@api_router.post("/reports")
async def create_report(report: ReportCreate, request: Request):
    user = await get_current_user(request)
    if report.target_type == "master_kit":
        target = await db.master_kits.find_one({"kit_id": report.target_id}, {"_id": 0})
    elif report.target_type == "version":
        target = await db.versions.find_one({"version_id": report.target_id}, {"_id": 0})
    else:
        raise HTTPException(status_code=400, detail="Invalid target type")
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    doc = {
        "report_id": f"rep_{uuid.uuid4().hex[:12]}",
        "target_type": report.target_type,
        "target_id": report.target_id,
        "original_data": target,
        "corrections": report.corrections,
        "notes": report.notes or "",
        "reported_by": user["user_id"],
        "reporter_name": user.get("name", ""),
        "status": "pending",
        "votes_up": 0,
        "votes_down": 0,
        "voters": [],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.reports.insert_one(doc)
    result = await db.reports.find_one({"report_id": doc["report_id"]}, {"_id": 0})
    return result

@api_router.get("/reports")
async def list_reports(status: Optional[str] = "pending", skip: int = 0, limit: int = 50):
    query = {}
    if status:
        query["status"] = status
    reports = await db.reports.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return reports

@api_router.post("/reports/{report_id}/vote")
async def vote_on_report(report_id: str, vote: VoteCreate, request: Request):
    user = await get_current_user(request)
    # Check user has at least 1 jersey in collection (moderators exempt)
    user_role = user.get("role", "user")
    is_moderator = user_role in ("moderator", "admin")
    if not is_moderator:
        col_count = await db.collections.count_documents({"user_id": user["user_id"]})
        if col_count == 0:
            raise HTTPException(status_code=403, detail="You must have at least 1 jersey in your collection to vote")
    report = await db.reports.find_one({"report_id": report_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    if report["status"] != "pending":
        raise HTTPException(status_code=400, detail="Report is no longer pending")
    if user["user_id"] in report.get("voters", []):
        raise HTTPException(status_code=400, detail="Already voted")
    if vote.vote not in ("up", "down"):
        raise HTTPException(status_code=400, detail="Vote must be 'up' or 'down'")
    
    # Moderators can approve with single vote
    vote_weight = APPROVAL_THRESHOLD if is_moderator and vote.vote == "up" else 1
    inc_field = "votes_up" if vote.vote == "up" else "votes_down"
    await db.reports.update_one(
        {"report_id": report_id},
        {"$inc": {inc_field: vote_weight}, "$push": {"voters": user["user_id"]}}
    )
    updated = await db.reports.find_one({"report_id": report_id}, {"_id": 0})
    if updated["votes_up"] >= APPROVAL_THRESHOLD:
        corrections = updated["corrections"]
        if updated["target_type"] == "master_kit":
            update_fields = {k: v for k, v in corrections.items() if k not in ("kit_id", "_id")}
            if update_fields:
                await db.master_kits.update_one({"kit_id": updated["target_id"]}, {"$set": update_fields})
        elif updated["target_type"] == "version":
            update_fields = {k: v for k, v in corrections.items() if k not in ("version_id", "_id")}
            if update_fields:
                await db.versions.update_one({"version_id": updated["target_id"]}, {"$set": update_fields})
        await db.reports.update_one({"report_id": report_id}, {"$set": {"status": "approved"}})
    return await db.reports.find_one({"report_id": report_id}, {"_id": 0})


# ─── Wishlist Routes ───

class WishlistAdd(BaseModel):
    version_id: str
    notes: Optional[str] = ""

@api_router.get("/wishlist")
async def get_wishlist(request: Request):
    user = await get_current_user(request)
    items = await db.wishlists.find({"user_id": user["user_id"]}, {"_id": 0}).sort("added_at", -1).to_list(500)
    enriched = []
    for item in items:
        version = await db.versions.find_one({"version_id": item["version_id"]}, {"_id": 0})
        if version:
            kit = await db.master_kits.find_one({"kit_id": version["kit_id"]}, {"_id": 0})
            item["version"] = version
            item["master_kit"] = kit
        enriched.append(item)
    return enriched

@api_router.post("/wishlist")
async def add_to_wishlist(item: WishlistAdd, request: Request):
    user = await get_current_user(request)
    version = await db.versions.find_one({"version_id": item.version_id}, {"_id": 0})
    if not version:
        raise HTTPException(status_code=404, detail="Version not found")
    existing = await db.wishlists.find_one(
        {"user_id": user["user_id"], "version_id": item.version_id}, {"_id": 0}
    )
    if existing:
        raise HTTPException(status_code=400, detail="Already in wishlist")
    doc = {
        "wishlist_id": f"wish_{uuid.uuid4().hex[:12]}",
        "user_id": user["user_id"],
        "version_id": item.version_id,
        "notes": item.notes or "",
        "added_at": datetime.now(timezone.utc).isoformat()
    }
    await db.wishlists.insert_one(doc)
    result = await db.wishlists.find_one({"wishlist_id": doc["wishlist_id"]}, {"_id": 0})
    return result

@api_router.delete("/wishlist/{wishlist_id}")
async def remove_from_wishlist(wishlist_id: str, request: Request):
    user = await get_current_user(request)
    result = await db.wishlists.delete_one({"wishlist_id": wishlist_id, "user_id": user["user_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Removed from wishlist"}

@api_router.get("/wishlist/check/{version_id}")
async def check_wishlist(version_id: str, request: Request):
    user = await get_current_user(request)
    existing = await db.wishlists.find_one(
        {"user_id": user["user_id"], "version_id": version_id}, {"_id": 0}
    )
    return {
        "in_wishlist": existing is not None,
        "wishlist_id": existing["wishlist_id"] if existing else None
    }


# ─── Autocomplete Route ───

@api_router.get("/autocomplete")
async def autocomplete(field: str, q: str = ""):
    field_map = {
        "club": "master_kits",
        "brand": "master_kits",
        "league": "master_kits",
        "sponsor": "master_kits",
        "competition": "versions",
    }
    if field not in field_map:
        return []
    collection_name = field_map[field]
    values = await db[collection_name].distinct(field)
    if q:
        q_lower = q.lower()
        values = [v for v in values if v and q_lower in str(v).lower()]
    return sorted([v for v in values if v])[:20]


# ─── Image Upload ───

ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

@api_router.post("/upload")
async def upload_image(file: UploadFile = File(...)):
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type {ext} not allowed. Use: {', '.join(ALLOWED_EXTENSIONS)}")
    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Max 10MB")
    filename = f"{uuid.uuid4().hex[:16]}{ext}"
    filepath = UPLOAD_DIR / filename
    async with aiofiles.open(filepath, "wb") as f:
        await f.write(contents)
    return {"filename": filename, "url": f"/api/uploads/{filename}"}

@api_router.post("/upload/multiple")
async def upload_multiple_images(files: List[UploadFile] = File(...)):
    results = []
    for file in files:
        ext = Path(file.filename).suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            continue
        filename = f"{uuid.uuid4().hex[:16]}{ext}"
        filepath = UPLOAD_DIR / filename
        async with aiofiles.open(filepath, "wb") as f:
            await f.write(contents)
        results.append({"filename": filename, "url": f"/api/uploads/{filename}"})
    return results


# ─── Seed / Import Data ───

@api_router.post("/seed")
async def seed_data():
    """Legacy seed endpoint - redirects to import"""
    return {"message": "Use POST /api/import-excel to import data from the Excel file"}

@api_router.post("/import-excel")
async def import_excel():
    """Clear DB and import master kits from the Excel file"""
    import openpyxl

    excel_path = Path("/tmp/Master_Kit_2005_2026.xlsx")
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Excel file not found at /tmp/Master_Kit_2005_2026.xlsx")

    # Clear existing data (keep users and sessions)
    for col_name in ["master_kits", "versions", "collections", "reviews", "reports", "submissions"]:
        await db[col_name].delete_many({})
    logger.info("Cleared existing data from master_kits, versions, collections, reviews, reports, submissions")

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    imported = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        col_map = {h: i for i, h in enumerate(headers) if h}

        for idx, row in enumerate(rows[1:]):
            start_year = 2005 + idx
            season = f"{start_year}/{start_year + 1}"

            team = row[col_map.get('Team', 0)] or ''
            if not team or str(team).strip() in ('', 'None'):
                continue  # Skip empty rows
            kit_type = row[col_map['Type']] if 'Type' in col_map else 'Home'
            design = row[col_map['Design']] if 'Design' in col_map and row[col_map['Design']] else ''
            colors = row[col_map['Colors']] if 'Colors' in col_map and row[col_map['Colors']] else ''
            brand = row[col_map['Brand']] if 'Brand' in col_map and row[col_map['Brand']] else ''
            sponsor = row[col_map['Sponsor (primary)']] if 'Sponsor (primary)' in col_map and row[col_map['Sponsor (primary)']] else ''
            league = row[col_map['League']] if 'League' in col_map and row[col_map['League']] else ''
            competition = row[col_map['Competition']] if 'Competition' in col_map and row[col_map['Competition']] else ''
            source_url = row[col_map['URL']] if 'URL' in col_map and row[col_map['URL']] else ''
            front_photo = row[col_map['Image URL']] if 'Image URL' in col_map and row[col_map['Image URL']] else ''

            doc = {
                "kit_id": f"kit_{uuid.uuid4().hex[:12]}",
                "club": str(team).strip(),
                "season": season,
                "kit_type": str(kit_type).strip() if kit_type else "Home",
                "brand": str(brand).strip() if brand else "",
                "front_photo": str(front_photo).strip() if front_photo else "",
                "year": start_year,
                "design": str(design).strip() if design else "",
                "colors": str(colors).strip() if colors else "",
                "sponsor": str(sponsor).strip() if sponsor else "",
                "league": str(league).strip() if league else "",
                "competition": str(competition).strip() if competition else "",
                "source_url": str(source_url).strip() if source_url else "",
                "created_by": "import",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.master_kits.insert_one(doc)
            imported += 1

    wb.close()
    logger.info(f"Imported {imported} master kits from Excel")
    return {"message": f"Successfully imported {imported} master kits", "count": imported}


@api_router.post("/migrate-schema")
async def migrate_schema():
    """Clean master_kits to match new schema: remove year, colors, competition, source_url fields"""
    # Fields to remove
    remove_fields = {"year", "colors", "competition", "source_url"}

    # Remove deprecated fields from all master_kits
    result = await db.master_kits.update_many(
        {},
        {"$unset": {f: "" for f in remove_fields}}
    )
    updated = result.modified_count

    # Ensure all docs have required fields with defaults
    all_kits = await db.master_kits.find({}, {"_id": 0}).to_list(1000)
    patched = 0
    for kit in all_kits:
        patch = {}
        if not kit.get("league"):
            patch["league"] = ""
        if not kit.get("design"):
            patch["design"] = ""
        if not kit.get("sponsor"):
            patch["sponsor"] = ""
        if not kit.get("gender"):
            patch["gender"] = ""
        if patch:
            await db.master_kits.update_one({"kit_id": kit["kit_id"]}, {"$set": patch})
            patched += 1

    # Also remove gender from versions (moved to master_kit level)
    ver_result = await db.versions.update_many({}, {"$unset": {"gender": ""}})

    return {
        "message": "Schema migration complete",
        "master_kits_cleaned": updated,
        "master_kits_patched": patched,
        "versions_cleaned": ver_result.modified_count
    }


@api_router.post("/migrate-create-default-versions")
async def migrate_create_default_versions():
    """Create default versions for all existing master kits that don't have any versions"""
    all_kits = await db.master_kits.find({}, {"_id": 0, "kit_id": 1, "front_photo": 1}).to_list(2000)
    
    created_count = 0
    skipped_count = 0
    
    for kit in all_kits:
        kit_id = kit.get("kit_id")
        if not kit_id:
            continue
            
        # Check if this kit already has any versions
        existing_version = await db.versions.find_one({"kit_id": kit_id}, {"_id": 1})
        if existing_version:
            skipped_count += 1
            continue
        
        # Create default version
        default_version = {
            "version_id": f"ver_{uuid.uuid4().hex[:12]}",
            "kit_id": kit_id,
            "competition": "National Championship",
            "model": "Replica",
            "sku_code": "",
            "ean_code": "",
            "front_photo": kit.get("front_photo", ""),
            "back_photo": "",
            "created_by": "system_migration",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.versions.insert_one(default_version)
        created_count += 1
    
    return {
        "message": "Default versions migration complete",
        "versions_created": created_count,
        "kits_skipped": skipped_count,
        "total_kits_processed": len(all_kits)
    }


@api_router.post("/set-moderator-role")
async def set_moderator_role():
    """Ensure moderator role is set for designated moderator emails"""
    updated_count = 0
    for email in MODERATOR_EMAILS:
        result = await db.users.update_one(
            {"email": email},
            {"$set": {"role": "moderator"}}
        )
        if result.modified_count > 0:
            updated_count += 1
    
    # Get current status
    moderators = await db.users.find(
        {"email": {"$in": MODERATOR_EMAILS}},
        {"_id": 0, "email": 1, "role": 1, "name": 1}
    ).to_list(100)
    
    return {
        "message": "Moderator roles updated",
        "updated_count": updated_count,
        "moderators": moderators
    }


# ─── Image Proxy ───

@api_router.get("/image-proxy")
async def image_proxy(url: str):
    """Proxy external images to avoid CORS/referrer issues"""
    if not url.startswith("https://cdn.footballkitarchive.com/"):
        raise HTTPException(status_code=400, detail="Only footballkitarchive CDN URLs allowed")
    async with httpx.AsyncClient() as hc:
        resp = await hc.get(url, timeout=10)
        if resp.status_code != 200:
            raise HTTPException(status_code=resp.status_code, detail="Failed to fetch image")
        content_type = resp.headers.get("content-type", "image/jpeg")
        return Response(content=resp.content, media_type=content_type, headers={"Cache-Control": "public, max-age=86400"})


# ─── Static File Serving ───

from fastapi.staticfiles import StaticFiles
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
