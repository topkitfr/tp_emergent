from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from typing import Optional
from datetime import datetime, timezone
from pathlib import Path
import uuid
import logging
import csv
import io
import re
from database import db
from models import ProfileUpdate
from auth import get_current_user
from utils import slugify, MODERATOR_EMAILS

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["admin"])


# ─── Helpers CSV ──────────────────────────────────────────────────────────────

def _csv_slug(text: str) -> str:
    """Slug stable : minuscules, alphanum + tirets, sans accents simples."""
    return re.sub(r'[^a-z0-9]+', '-', text.lower().strip()).strip('-')


def _normalize_season(season: str) -> str:
    """
    Normalise n'importe quel format de saison vers 'YYYY/YYYY+1'.
    Gère : '00-01', '00-01 (Carry-over)', '2022-23', '1987-88',
            '01/01/2000' (date parasite → skip via return ''),
            '01-03' (span multi-années → skip via return '')
    """
    if not season:
        return ""
    s = re.sub(r'\s*\(carry-over\)', '', season, flags=re.IGNORECASE).strip()

    # Format dd/mm/yyyy ou similaire → inutilisable comme saison
    if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', s):
        return ""

    # Format standard : YY-YY ou YYYY-YY ou YYYY-YYYY
    m = re.match(r'^(\d{2,4})[-/](\d{2,4})$', s)
    if m:
        y1_str, y2_str = m.group(1), m.group(2)

        # Résoudre y1
        if len(y1_str) == 2:
            n = int(y1_str)
            y1 = 1900 + n if n >= 85 else 2000 + n
        else:
            y1 = int(y1_str)

        # Résoudre y2
        if len(y2_str) == 2:
            n2 = int(y2_str)
            # y2 doit être y1+1 (ex: 00-01 → 2001, pas 2000+0=2000)
            # On vérifie la cohérence
            y2_candidate = 1900 + n2 if n2 >= 85 else 2000 + n2
            # Si l'écart est > 2 ans c'est un span multi-saisons → skip
            if abs(y2_candidate - y1) > 2:
                return ""
            y2 = y1 + 1  # toujours normaliser à y1/y1+1
        else:
            y2 = int(y2_str)
            if abs(y2 - y1) > 2:
                return ""
            y2 = y1 + 1

        return f"{y1}/{y2}"

    return ""  # Format non reconnu → skip


def _normalize_league(league: str) -> str:
    """
    Normalise les noms de ligues : corrige les slugs et les variantes.
    """
    mapping = {
        'la-liga':              'La Liga',
        'bundesliga':           'Bundesliga',
        'ligue-1':              'Ligue 1',
        'premier-league':       'Premier League',
        'serie-a':              'Serie A',
        # Variantes historiques → noms modernes
        'championnat de france': 'Ligue 1',
        'division 1':            'Ligue 1',
    }
    if not league:
        return ''
    normalized = mapping.get(league.lower().strip(), league.strip())
    return normalized


def _normalize_sponsor(sponsor: str) -> str:
    """Nettoie les sponsors parasites."""
    s = sponsor.strip() if sponsor else ''
    return '' if s in ('-', 'N/A', 'n/a', 'None') else s


def _normalize_gender(gender: str) -> str:
    g = (gender or '').strip().upper()
    mapping = {
        'MAN': 'MEN', 'MALE': 'MEN',
        'WOMAN': 'WOMEN', 'FEMALE': 'WOMEN',
        'KID': 'YOUTH', 'KIDS': 'YOUTH', 'CHILD': 'YOUTH',
        '': 'MEN',
    }
    return g if g in ('MEN', 'WOMEN', 'YOUTH', 'UNISEX') else mapping.get(g, 'MEN')


# ─── Stats ────────────────────────────────────────────────────────────────────

@router.get("/stats")
async def get_stats():
    kits     = await db.master_kits.count_documents({})
    versions = await db.versions.count_documents({})
    users    = await db.users.count_documents({})
    reviews  = await db.reviews.count_documents({})
    return {"master_kits": kits, "versions": versions, "users": users, "reviews": reviews}


# ─── User Profile ─────────────────────────────────────────────────────────────

@router.get("/users/by-username/{username}")
async def get_user_by_username(username: str):
    user = await db.users.find_one({"username": username}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user_id = user["user_id"]
    user["collection_count"] = await db.collections.count_documents({"user_id": user_id})
    user["review_count"]     = await db.reviews.count_documents({"user_id": user_id})
    user["submission_count"] = await db.submissions.count_documents({"submitted_by": user_id})
    return user


@router.get("/users/{user_id}/profile")
async def get_user_profile(user_id: str):
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user["collection_count"] = await db.collections.count_documents({"user_id": user_id})
    user["review_count"]     = await db.reviews.count_documents({"user_id": user_id})
    user["submission_count"] = await db.submissions.count_documents({"submitted_by": user_id})
    return user


@router.put("/users/profile")
async def update_profile(update: ProfileUpdate, request: Request):
    user = await get_current_user(request)
    update_dict = {k: v for k, v in update.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    if "username" in update_dict:
        existing = await db.users.find_one(
            {"username": update_dict["username"], "user_id": {"$ne": user["user_id"]}}, {"_id": 0}
        )
        if existing:
            raise HTTPException(status_code=400, detail="Username already taken")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": update_dict})
    return await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})


# ─── Seed / Import Data ───────────────────────────────────────────────────────

@router.post("/seed")
async def seed_data():
    return {"message": "Use POST /api/import-excel to import data from the Excel file"}


@router.post("/import-excel")
async def import_excel():
    import openpyxl

    excel_path = Path("/tmp/Master_Kit_2005_2026.xlsx")
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Excel file not found at /tmp/Master_Kit_2005_2026.xlsx")

    for col_name in ["master_kits", "versions", "collections", "reviews", "reports", "submissions"]:
        await db[col_name].delete_many({})
    logger.info("Cleared existing data")

    wb   = openpyxl.load_workbook(excel_path, read_only=True)
    imported = 0

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]
        col_map = {h: i for i, h in enumerate(headers) if h}

        for idx, row in enumerate(rows[1:]):
            start_year = 2005 + idx
            season = f"{start_year}/{start_year + 1}"
            team = row[col_map.get('Team', 0)] or ''
            if not team or str(team).strip() in ('', 'None'):
                continue
            kit_type   = row[col_map['Type']]               if 'Type'             in col_map and row[col_map['Type']]             else 'Home'
            design     = row[col_map['Design']]             if 'Design'           in col_map and row[col_map['Design']]           else ''
            colors     = row[col_map['Colors']]             if 'Colors'           in col_map and row[col_map['Colors']]           else ''
            brand      = row[col_map['Brand']]              if 'Brand'            in col_map and row[col_map['Brand']]            else ''
            sponsor    = row[col_map['Sponsor (primary)']]  if 'Sponsor (primary)' in col_map and row[col_map['Sponsor (primary)']] else ''
            league     = row[col_map['League']]             if 'League'           in col_map and row[col_map['League']]           else ''
            competition= row[col_map['Competition']]        if 'Competition'      in col_map and row[col_map['Competition']]      else ''
            source_url = row[col_map['URL']]                if 'URL'              in col_map and row[col_map['URL']]              else ''
            front_photo= row[col_map['Image URL']]          if 'Image URL'        in col_map and row[col_map['Image URL']]        else ''

            doc = {
                "kit_id": f"kit_{uuid.uuid4().hex[:12]}",
                "club": str(team).strip(),
                "season": season,
                "kit_type": str(kit_type).strip() if kit_type else "Home",
                "brand": str(brand).strip() if brand else "",
                "front_photo": str(front_photo).strip() if front_photo else "",
                "year": start_year,
                "design": str(design).strip() if design else "",
                "colors": str(colors).strip() if colors else "",
                "sponsor": str(sponsor).strip() if sponsor else "",
                "league": str(league).strip() if league else "",
                "competition": str(competition).strip() if competition else "",
                "source_url": str(source_url).strip() if source_url else "",
                "created_by": "import",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.master_kits.insert_one(doc)
            imported += 1

    wb.close()
    return {"message": f"Successfully imported {imported} master kits", "count": imported}


# ─── Migration Endpoints ──────────────────────────────────────────────────────

@router.post("/migrate-schema")
async def migrate_schema():
    remove_fields = {"year", "colors", "competition", "source_url"}
    result  = await db.master_kits.update_many({}, {"$unset": {f: "" for f in remove_fields}})
    updated = result.modified_count
    all_kits = await db.master_kits.find({}, {"_id": 0}).to_list(1000)
    patched = 0
    for kit in all_kits:
        patch = {}
        for field in ("league", "design", "sponsor", "gender"):
            if not kit.get(field):
                patch[field] = ""
        if patch:
            await db.master_kits.update_one({"kit_id": kit["kit_id"]}, {"$set": patch})
            patched += 1
    ver_result = await db.versions.update_many({}, {"$unset": {"gender": ""}})
    return {
        "message": "Schema migration complete",
        "master_kits_cleaned": updated,
        "master_kits_patched": patched,
        "versions_cleaned": ver_result.modified_count
    }


@router.post("/migrate-create-default-versions")
async def migrate_create_default_versions():
    all_kits = await db.master_kits.find({}, {"_id": 0, "kit_id": 1, "front_photo": 1}).to_list(2000)
    created_count = skipped_count = 0
    for kit in all_kits:
        kit_id = kit.get("kit_id")
        if not kit_id:
            continue
        if await db.versions.find_one({"kit_id": kit_id}, {"_id": 1}):
            skipped_count += 1
            continue
        await db.versions.insert_one({
            "version_id":  f"ver_{uuid.uuid4().hex[:12]}",
            "kit_id":      kit_id,
            "competition": "National Championship",
            "model":       "Replica",
            "sku_code":    "",
            "ean_code":    "",
            "front_photo": kit.get("front_photo", ""),
            "back_photo":  "",
            "created_by":  "system_migration",
            "created_at":  datetime.now(timezone.utc).isoformat()
        })
        created_count += 1
    return {
        "message": "Default versions migration complete",
        "versions_created": created_count,
        "kits_skipped": skipped_count,
        "total_kits_processed": len(all_kits)
    }


@router.post("/set-moderator-role")
async def set_moderator_role():
    updated_count = 0
    for email in MODERATOR_EMAILS:
        result = await db.users.update_one({"email": email}, {"$set": {"role": "moderator"}})
        if result.modified_count > 0:
            updated_count += 1
    moderators = await db.users.find(
        {"email": {"$in": MODERATOR_EMAILS}},
        {"_id": 0, "email": 1, "role": 1, "name": 1}
    ).to_list(100)
    return {"message": "Moderator roles updated", "updated_count": updated_count, "moderators": moderators}


@router.post("/migrate-entities-from-kits")
async def migrate_entities_from_kits():
    all_kits = await db.master_kits.find({}, {"_id": 0}).to_list(5000)
    now = datetime.now(timezone.utc).isoformat()
    teams_created = leagues_created = brands_created = kits_updated = 0

    async def _upsert_entity(collection, id_field, slug_val, doc):
        existing = await db[collection].find_one({"slug": slug_val}, {"_id": 0, id_field: 1})
        if existing:
            return existing[id_field]
        await db[collection].insert_one(doc)
        return doc[id_field]

    team_map, league_map, brand_map = {}, {}, {}

    for name in sorted(set(k.get("club", "").strip() for k in all_kits if k.get("club", "").strip())):
        sl = slugify(name)
        tid = await _upsert_entity("teams", "team_id", sl, {
            "team_id": f"team_{uuid.uuid4().hex[:12]}", "name": name, "slug": sl,
            "country": "", "city": "", "founded": None, "primary_color": "",
            "secondary_color": "", "crest_url": "", "aka": [], "created_at": now, "updated_at": now
        })
        team_map[name] = tid
        if tid not in team_map.values():
            teams_created += 1

    for name in sorted(set(k.get("league", "").strip() for k in all_kits if k.get("league", "").strip())):
        sl = slugify(name)
        lid = await _upsert_entity("leagues", "league_id", sl, {
            "league_id": f"league_{uuid.uuid4().hex[:12]}", "name": name, "slug": sl,
            "country_or_region": "", "level": "domestic", "organizer": "", "logo_url": "",
            "kit_count": 0, "status": "approved", "created_at": now, "updated_at": now
        })
        league_map[name] = lid

    for name in sorted(set(k.get("brand", "").strip() for k in all_kits if k.get("brand", "").strip())):
        sl = slugify(name)
        bid = await _upsert_entity("brands", "brand_id", sl, {
            "brand_id": f"brand_{uuid.uuid4().hex[:12]}", "name": name, "slug": sl,
            "country": "", "founded": None, "logo_url": "", "kit_count": 0,
            "status": "approved", "created_at": now, "updated_at": now
        })
        brand_map[name] = bid

    for kit in all_kits:
        update = {}
        if kit.get("club", "").strip() in team_map:
            update["team_id"] = team_map[kit["club"].strip()]
        if kit.get("league", "").strip() in league_map:
            update["league_id"] = league_map[kit["league"].strip()]
        if kit.get("brand", "").strip() in brand_map:
            update["brand_id"] = brand_map[kit["brand"].strip()]
        if update:
            await db.master_kits.update_one({"kit_id": kit["kit_id"]}, {"$set": update})
            kits_updated += 1

    return {
        "message": "Entity migration complete",
        "teams_created": teams_created, "leagues_created": leagues_created,
        "brands_created": brands_created, "kits_updated": kits_updated,
    }


# ─── Reset + Import CSV (admin/moderator only) ────────────────────────────────

@router.post("/admin/reset-and-import-csv")
async def reset_and_import_csv(request: Request, file: UploadFile = File(...)):
    """
    1. Vérifie le rôle admin/moderator.
    2. Vide les collections : master_kits, versions, teams, leagues, brands.
       (users, collections, reviews, wishlist, submissions sont préservés)
    3. Importe le CSV avec normalisation complète :
       - Saisons → YYYY/YYYY+1 (multi-années et dates invalides skippées)
       - Ligues  → noms canoniques (la-liga → La Liga, etc.)
       - Sponsors '-' / '' → ''
       - Doublons (team+season+type) ignorés
       - Entités teams/brands/leagues créées ou réutilisées via slug
    4. Crée une version par défaut (Replica / National Championship) pour chaque kit.
    """
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "moderator"):
        raise HTTPException(status_code=403, detail="Rôle admin ou modérateur requis")

    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Le fichier doit être un .csv")

    # ── 1. Lecture du CSV ──────────────────────────────────────────────────────
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    rows   = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail="CSV vide ou mal formaté")

    # ── 2. Reset des collections kits/entités ──────────────────────────────────
    for col in ("master_kits", "versions", "teams", "leagues", "brands"):
        await db[col].delete_many({})
    logger.info(f"reset-and-import-csv: reset par {user.get('email')} — {len(rows)} lignes CSV")

    now = datetime.now(timezone.utc).isoformat()

    # ── 3. Pré-construire les maps entités (batch pour éviter N×get_or_create) ─
    # Dédoublonner d'abord les données utiles
    seen_kits: set[str] = set()
    valid_rows = []
    for r in rows:
        team_name = (r.get("team") or "").strip()
        season_raw = (r.get("season") or "").strip()
        kit_type   = (r.get("type") or "Home").strip() or "Home"
        if not team_name or not season_raw:
            continue
        season = _normalize_season(season_raw)
        if not season:
            continue
        key = f"{team_name.lower()}|{season}|{kit_type.lower()}"
        if key in seen_kits:
            continue
        seen_kits.add(key)
        valid_rows.append({
            "team":       team_name,
            "season":     season,
            "kit_type":   kit_type,
            "design":     (r.get("design") or "").strip(),
            "colors":     (r.get("colors") or "").strip(),
            "brand":      (r.get("brand") or "").strip(),
            "sponsor":    _normalize_sponsor(r.get("sponsor") or ""),
            "league":     _normalize_league(r.get("league") or ""),
            "gender":     _normalize_gender(r.get("gender") or ""),
            "img_url":    (r.get("img_url") or r.get("Image URL") or "").strip(),
            "source_url": (r.get("source_url") or r.get("URL") or "").strip(),
        })

    logger.info(f"reset-and-import-csv: {len(valid_rows)} lignes valides après dédup/normalisation (sur {len(rows)})")

    # ── 4. Créer les entités en batch (1 INSERT par entité unique) ─────────────
    all_team_names   = sorted(set(r["team"]   for r in valid_rows if r["team"]))
    all_league_names = sorted(set(r["league"] for r in valid_rows if r["league"]))
    all_brand_names  = sorted(set(r["brand"]  for r in valid_rows if r["brand"]))

    team_map:   dict[str, str] = {}
    league_map: dict[str, str] = {}
    brand_map:  dict[str, str] = {}

    for name in all_team_names:
        team_id = f"team_{uuid.uuid4().hex[:12]}"
        sl = slugify(name)
        await db.teams.insert_one({
            "team_id": team_id, "name": name, "slug": sl,
            "country": "", "city": "", "founded": None,
            "primary_color": "", "secondary_color": "",
            "crest_url": "", "aka": [], "kit_count": 0,
            "status": "approved", "created_at": now, "updated_at": now,
        })
        team_map[name] = team_id

    for name in all_league_names:
        league_id = f"league_{uuid.uuid4().hex[:12]}"
        sl = slugify(name)
        await db.leagues.insert_one({
            "league_id": league_id, "name": name, "slug": sl,
            "country_or_region": "", "level": "domestic",
            "organizer": "", "logo_url": "", "kit_count": 0,
            "status": "approved", "created_at": now, "updated_at": now,
        })
        league_map[name] = league_id

    for name in all_brand_names:
        brand_id = f"brand_{uuid.uuid4().hex[:12]}"
        sl = slugify(name)
        await db.brands.insert_one({
            "brand_id": brand_id, "name": name, "slug": sl,
            "country": "", "founded": None, "logo_url": "",
            "kit_count": 0, "status": "approved",
            "created_at": now, "updated_at": now,
        })
        brand_map[name] = brand_id

    # ── 5. Insérer les master_kits + une version par défaut ────────────────────
    kits_to_insert     = []
    versions_to_insert = []

    for r in valid_rows:
        kit_id   = f"kit_{uuid.uuid4().hex[:12]}"
        kit_slug = _csv_slug(f"{r['team']}-{r['season']}-{r['kit_type']}")
        team_id  = team_map.get(r["team"], "")
        league_id = league_map.get(r["league"], "")
        brand_id  = brand_map.get(r["brand"], "")

        kits_to_insert.append({
            "kit_id":      kit_id,
            "slug":        kit_slug,
            "team_id":     team_id,
            "brand_id":    brand_id,
            "league_id":   league_id,
            "club":        r["team"],
            "brand":       r["brand"],
            "league":      r["league"],
            "season":      r["season"],
            "kit_type":    r["kit_type"],
            "type":        r["kit_type"],
            "design":      r["design"],
            "colors":      r["colors"],
            "sponsor":     r["sponsor"],
            "gender":      r["gender"],
            "front_photo": r["img_url"],
            "img_url":     r["img_url"],
            "source_url":  r["source_url"],
            "status":      "approved",
            "avg_rating":  0.0,
            "version_count": 1,
            "created_by":  user.get("user_id", "admin"),
            "created_at":  now,
        })

        versions_to_insert.append({
            "version_id":  f"ver_{uuid.uuid4().hex[:12]}",
            "kit_id":      kit_id,
            "competition": "National Championship",
            "model":       "Replica",
            "sku_code":    "",
            "ean_code":    "",
            "front_photo": r["img_url"],
            "back_photo":  "",
            "main_player_id": "",
            "avg_rating":  0.0,
            "review_count": 0,
            "created_by":  user.get("user_id", "admin"),
            "created_at":  now,
        })

    # Insert en batch
    if kits_to_insert:
        await db.master_kits.insert_many(kits_to_insert, ordered=False)
    if versions_to_insert:
        await db.versions.insert_many(versions_to_insert, ordered=False)

    skipped = len(rows) - len(valid_rows)
    logger.info(
        f"reset-and-import-csv done: {len(kits_to_insert)} kits, "
        f"{len(versions_to_insert)} versions, {skipped} lignes skippées, "
        f"{len(team_map)} teams, {len(league_map)} leagues, {len(brand_map)} brands"
    )

    return {
        "message": "Reset et import terminés",
        "kits_created":     len(kits_to_insert),
        "versions_created": len(versions_to_insert),
        "teams_created":    len(team_map),
        "leagues_created":  len(league_map),
        "brands_created":   len(brand_map),
        "rows_total":       len(rows),
        "rows_skipped":     skipped,
        "skip_reasons": {
            "no_team_or_season": "Lignes sans team ou sans saison",
            "invalid_season":    "Saisons multi-années ou dates invalides ignorées",
            "duplicates":        "Doublons team+season+type ignorés",
        }
    }


# ─── Import CSV simple (sans reset, admin/moderator only) ────────────────────

@router.post("/admin/import-csv")
async def import_csv_upload(request: Request, file: UploadFile = File(...)):
    """
    Upload un CSV et importe les kits sans reset.
    Les doublons existants (même slug) sont ignorés.
    """
    user = await get_current_user(request)
    if user.get("role") not in ("admin", "moderator"):
        raise HTTPException(status_code=403, detail="Rôle admin ou modérateur requis")

    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Le fichier doit être un .csv")

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    rows   = list(reader)
    if not rows:
        raise HTTPException(status_code=400, detail="CSV vide ou mal formaté")

    now = datetime.now(timezone.utc).isoformat()
    created_kits = 0
    skipped_kits = 0
    import_errors: list[str] = []

    async def _get_or_create(collection, id_field, slug_val, doc):
        existing = await db[collection].find_one({"slug": slug_val}, {"_id": 0, id_field: 1})
        if existing:
            return existing[id_field]
        await db[collection].insert_one(doc)
        return doc[id_field]

    for i, r in enumerate(rows, 1):
        try:
            team_name  = (r.get("team") or "").strip()
            season_raw = (r.get("season") or "").strip()
            kit_type   = (r.get("type") or "Home").strip() or "Home"

            if not team_name or not season_raw:
                skipped_kits += 1
                continue
            season = _normalize_season(season_raw)
            if not season:
                skipped_kits += 1
                continue

            brand_name  = (r.get("brand") or "").strip()
            league_name = _normalize_league(r.get("league") or "")
            img_url     = (r.get("img_url") or r.get("Image URL") or "").strip()

            # Skip doublon
            kit_slug = _csv_slug(f"{team_name}-{season}-{kit_type}")
            if await db["master_kits"].find_one({"slug": kit_slug}):
                skipped_kits += 1
                continue

            team_id = await _get_or_create("teams", "team_id", slugify(team_name), {
                "team_id": f"team_{uuid.uuid4().hex[:12]}", "name": team_name, "slug": slugify(team_name),
                "country": "", "city": "", "founded": None, "primary_color": "", "secondary_color": "",
                "crest_url": "", "aka": [], "kit_count": 0, "status": "approved",
                "created_at": now, "updated_at": now,
            })
            brand_id = ""
            if brand_name:
                brand_id = await _get_or_create("brands", "brand_id", slugify(brand_name), {
                    "brand_id": f"brand_{uuid.uuid4().hex[:12]}", "name": brand_name, "slug": slugify(brand_name),
                    "country": "", "founded": None, "logo_url": "", "kit_count": 0,
                    "status": "approved", "created_at": now, "updated_at": now,
                })
            league_id = ""
            if league_name:
                league_id = await _get_or_create("leagues", "league_id", slugify(league_name), {
                    "league_id": f"league_{uuid.uuid4().hex[:12]}", "name": league_name, "slug": slugify(league_name),
                    "country_or_region": "", "level": "domestic", "organizer": "", "logo_url": "",
                    "kit_count": 0, "status": "approved", "created_at": now, "updated_at": now,
                })

            kit_id = f"kit_{uuid.uuid4().hex[:12]}"
            await db["master_kits"].insert_one({
                "kit_id": kit_id, "slug": kit_slug,
                "team_id": team_id, "brand_id": brand_id, "league_id": league_id,
                "club": team_name, "brand": brand_name, "league": league_name,
                "season": season, "kit_type": kit_type, "type": kit_type,
                "design":      (r.get("design") or "").strip(),
                "colors":      (r.get("colors") or "").strip(),
                "sponsor":     _normalize_sponsor(r.get("sponsor") or ""),
                "gender":      _normalize_gender(r.get("gender") or ""),
                "front_photo": img_url, "img_url": img_url,
                "source_url":  (r.get("source_url") or "").strip(),
                "status": "approved", "avg_rating": 0.0,
                "created_by": user.get("user_id", "admin"),
                "created_at": now,
            })
            created_kits += 1

        except Exception as e:
            if len(import_errors) < 20:
                import_errors.append(f"Ligne {i} ({r.get('team','?')}): {type(e).__name__}: {e}")

    return {
        "message": "Import terminé",
        "created":    created_kits,
        "skipped":    skipped_kits,
        "total_rows": len(rows),
        "errors":     import_errors,
    }
